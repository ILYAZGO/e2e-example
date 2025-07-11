#from playwright.sync_api import Page, expect, BrowserContext, sync_playwright, Playwright, Route
from utils.variables import *
from pages.communications import *
from utils.dates import *
from datetime import datetime
from utils.create_delete_user import create_user, delete_user, give_access_right, give_users_to_manager
import os
import pytest
import allure
from openpyxl import load_workbook


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_dates")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("Check dates buttons")
def test_check_dates(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with Ecotelecom"):
        communications.auth(ECOTELECOM, ECOPASS)

    with allure.step("Check first and last dates in view. Today by default"):
        communications.assert_check_period_dates(today.strftime("%d/%m/%Y"), today.strftime("%d/%m/%Y"))

    with allure.step("Switch to yesterday"):
        communications.yesterday.click()

    with allure.step("Check first and last dates in view."):
        communications.assert_check_period_dates(yesterday.strftime("%d/%m/%Y"), yesterday.strftime("%d/%m/%Y"))

    with allure.step("Click to week"):
        communications.week.click()

    with allure.step("Choose this week"):
        communications.select_period_value("this_week")

    with allure.step("Check first and last dates in view."):
        communications.assert_check_period_dates(first_day_this_week.strftime("%d/%m/%Y"), last_day_this_week.strftime("%d/%m/%Y"))

    with allure.step("Click to week"):
        communications.week.click()

    with allure.step("Choose last week"):
        communications.select_period_value("last_week")

    with allure.step("Check first and last dates in view."):
        communications.assert_check_period_dates(first_day_last_week.strftime("%d/%m/%Y"), last_day_last_week.strftime("%d/%m/%Y"))

    with allure.step("Click to month"):
        communications.month.click()

    with allure.step("Choose this month"):
        communications.select_period_value("this_month")

    with allure.step("Check first and last dates in view."):
        communications.assert_check_period_dates(first_day_this_month.strftime("%d/%m/%Y"), last_day_this_month.strftime("%d/%m/%Y"))

    with allure.step("Click to month"):
        communications.month.click()

    with allure.step("Choose last month"):
        communications.select_period_value("last_month")

    with allure.step("Check first and last dates in view."):
        communications.assert_check_period_dates(first_day_last_month.strftime("%d/%m/%Y"), last_day_last_month.strftime("%d/%m/%Y"))

    with allure.step("Click to quarter"):
        communications.quarter.click()

    with allure.step("Choose this quarter"):
        communications.select_period_value("this_quarter")

    with allure.step("Check first and last dates in view."):
        communications.assert_check_period_dates(first_day_this_quarter.strftime("%d/%m/%Y"), last_day_this_quarter.strftime("%d/%m/%Y"))

    with allure.step("Click to quarter"):
        communications.quarter.click()

    with allure.step("Choose last quarter"):
        communications.select_period_value("last_quarter")

    # with allure.step("Check first and last dates in view."):
    #     communications.assert_check_period_dates(first_day_last_quarter.strftime("%d/%m/%Y"), last_day_last_quarter.strftime("%d/%m/%Y"))

    with allure.step("Click to year"):
        communications.year.click()

    with allure.step("Choose this year"):
        communications.select_period_value("this_year")

    with allure.step("Check first and last dates in view."):
        communications.assert_check_period_dates(first_day_this_year.strftime("%d/%m/%Y"), last_day_this_year.strftime("%d/%m/%Y"))

    with allure.step("Click to year"):
        communications.year.click()

    with allure.step("Choose last year"):
        communications.select_period_value("last_year")

    with allure.step("Check first and last dates in view."):
        communications.assert_check_period_dates(first_day_last_year.strftime("%d/%m/%Y"), last_day_last_year.strftime("%d/%m/%Y"))

    with allure.step("Switch to all time"):
        communications.all_time.click()

    with allure.step("Check begin and end dates is disabled"):
        communications.assert_check_period_dates_disabled()


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_communications_check_calendar_localization")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_communications_check_calendar_localization")
def test_communications_check_calendar_localization(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN, PASSWORD)

    with allure.step("Click to calendar"):
        page.wait_for_timeout(3000)
        communications.first_date.click()

    with allure.step("Check localization"):
        expect(page.locator('[class="ant-picker-content"]').nth(0)).to_contain_text("ПнВтСрЧтПтСбВс")
        expect(page.locator('[class="ant-picker-content"]').nth(1)).to_contain_text("ПнВтСрЧтПтСбВс")

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_search_all")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Searching all communications for Ecotelecom")
def test_check_search_all(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with Ecotelecom"):
        communications.auth(ECOTELECOM, ECOPASS)

    with allure.step("Choose period from 01/01/2022 to 31/12/2022"):
        communications.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_more_than_50()

    with allure.step("Check that all communications found"):
        communications.assert_communications_found("Найдено коммуникаций 3130 из 3130")

    with allure.step("Check that 50 calls in one page"):
        expect(communications.button_share_call).to_have_count(50)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_search_by_client_number")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check search by client number for Ecotelecom")
def test_check_search_by_client_number(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with Ecotelecom"):
        communications.auth(ECOTELECOM, ECOPASS)

    with allure.step("Choose period from 01/01/2022 to 31/12/2022"):
        communications.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Fill client number"):
        communications.fill_client_number("79251579005")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_less_than_50()

    with allure.step("Check"):
        communications.assert_communications_found("Найдено коммуникаций 14 из 3130")



@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_search_by_employee_number")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check search by employee's number for Ecotelecom")
def test_check_search_by_employee_number(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with Ecotelecom"):
        communications.auth(ECOTELECOM, ECOPASS)

    with allure.step("Choose period from 01/01/2022 to 31/12/2022"):
        communications.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Fill employee number"):
        communications.fill_employee_number("4995055555")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_more_than_50()

    with allure.step("Check"):
        communications.assert_communications_found("Найдено коммуникаций 670 из 3130")


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_search_by_client_dict_or_text")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check search by client dict or text for Ecotelecom")
def test_check_search_by_client_dict_or_text(base_url, page: Page) -> None:

    communications = Communications(page)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with Ecotelecom"):
        communications.auth(ECOTELECOM, ECOPASS)

    with allure.step("Choose period from 01/01/2022 to 31/12/2022"):
        communications.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Fill input with text"):
        communications.fill_client_dict_or_text("минутку", "минутку")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_less_than_50()

    with allure.step("Check"):
        communications.assert_communications_found("Найдено коммуникаций 20 из 3130")

    with allure.step("Clear, fill input by dict, choose dict from suggestion"):
        communications.fill_client_dict_or_text("Зо", "Словарь: Зомбоящик")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_more_than_50()

    with allure.step("Check"):
        communications.assert_communications_found("Найдено коммуникаций 405 из 3130")


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_search_by_employee_dict_or_text")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check search by employee's dict or text for Ecotelecom")
def test_check_search_by_employee_dict_or_text(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with Ecotelecom"):
        communications.auth(ECOTELECOM, ECOPASS)

    with allure.step("Choose period from 01/01/2022 to 31/12/2022"):
        communications.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Fill input with text"):
        communications.fill_employee_dict_or_text("минутку", "минутку")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_more_than_50()

    with allure.step("Check"):
        communications.assert_communications_found("Найдено коммуникаций 204 из 3130")

    with allure.step("Clear, fill input by dict, choose dict from suggestion"):
        communications.fill_employee_dict_or_text("Зо", "Словарь: Зомбоящик")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_more_than_50()

    with allure.step("Check"):
        communications.assert_communications_found("Найдено коммуникаций 492 из 3130")


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_search_by_exact_time")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check search by exact time for Ecotelecom")
def test_check_search_by_exact_time(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with Ecotelecom"):
        communications.auth(ECOTELECOM, ECOPASS)

    with allure.step("Choose period from 01/01/2022 to 31/12/2022"):
        communications.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Fill exact time"):
        communications.fill_time("11:42")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_less_than_50()

    with allure.step("Check"):
        communications.assert_communications_found("Найдено коммуникаций 5 из 3130")


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_search_by_length")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check search by length for Ecotelecom")
def test_check_search_by_length(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with Ecotelecom"):
        communications.auth(ECOTELECOM, ECOPASS)

    with allure.step("Choose period from 01/01/2022 to 31/12/2022"):
        communications.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Fill length <10"):
        communications.fill_search_length("<10")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_more_than_50()

    with allure.step("Check"):
        communications.assert_communications_found("Найдено коммуникаций 441 из 3130")

    with allure.step("Fill length >10"):
        communications.fill_search_length(">10")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_more_than_50()

    with allure.step("Check"):
        communications.assert_communications_found("Найдено коммуникаций 2689 из 3130")

    with allure.step("Fill length 1711"):
        communications.fill_search_length("1711")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_less_than_50()

    with allure.step("Check"):
        communications.assert_communications_found("Найдено коммуникаций 1 из 3130")


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_search_by_id")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check search by id for Ecotelecom")
def test_check_search_by_id(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with Ecotelecom"):
        communications.auth(ECOTELECOM, ECOPASS)

    with allure.step("Choose period from 01/01/2022 to 31/12/2022"):
        communications.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Fill ID"):
        communications.fill_id("1644268426.90181")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_less_than_50()

    with allure.step("Check that commincation found"):
        communications.assert_communications_found("Найдено коммуникаций 1 из 3130")

    with allure.step("Check extra"):
        expect(page.get_by_text("Теги сделки")).to_have_count(1)
        expect(page.get_by_text("Теги коммуникации")).to_have_count(1)
        expect(page.get_by_text("Теги фрагментов")).to_have_count(1)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_search_by_tag")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check search by tag for Ecotelecom")
def test_check_search_by_tag(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with Ecotelecom"):
        communications.auth(ECOTELECOM, ECOPASS)

    with allure.step("Choose period from 01/01/2022 to 31/12/2022"):
        communications.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Fill by tag"):
        communications.fill_by_tag(0, "Другой отдел")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_less_than_50()

    with allure.step("Check"):
        communications.assert_communications_found("Найдено коммуникаций 131 из 3130")

    with allure.step("Add extra tag"):
        communications.fill_by_tag(0, "Обсуждение тарифа")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_less_than_50()

    with allure.step("Check"):
        communications.assert_communications_found("Найдено коммуникаций 46 из 3130")

    with allure.step("Click to (Add condition)"):
        page.locator(BUTTON_ADD_CONDITION).first.click()

    with allure.step("Change logic operator"):
        page.locator('[data-testid="filters_search_by_tags"]').nth(1).get_by_text("ИЛИ").click()
        page.wait_for_selector(MENU)
        page.get_by_text("НЕТ ВСЕХ").click()

    with allure.step("Add new tag"):
        communications.fill_by_tag(1, "Новое подключение")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_less_than_50()

    with allure.step("Check"):
        communications.assert_communications_found("Найдено коммуникаций 19 из 3130")


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_sort")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("Check sort (6 type) all calls for Ecotelecom")
def test_check_sort(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with Ecotelecom"):
        communications.auth(ECOTELECOM, ECOPASS)

    with allure.step("Choose period from 01/01/2022 to 31/12/2022"):
        communications.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_more_than_50()

    with allure.step("Check all communications count, OLD calls first by default"):
        communications.assert_communications_found("Найдено коммуникаций 3130 из 3130")
        communications.assert_call_date_and_time("08.02.22 00:12")

    with allure.step("Change sort to NEW FIRST"):
        communications.change_sort("Сначала новые")

    with allure.step("Check NEW FIRST in list"):
        communications.assert_call_date_and_time("16.05.22 18:21")

    with allure.step("Change sort to SHORT FIRST"):
        communications.change_sort("Сначала короткие")

    with allure.step("Check SHORT FIRST in list"):
        communications.assert_call_date_and_time("10.02.22 09:49")  #09.02.22 11:41   "09.02.22 16:50"

    with allure.step("Change sort to LONG FIRST"):
        communications.change_sort("Сначала длинные")

    with allure.step("Check LONG FIRST in list"):
        communications.assert_call_date_and_time("09.02.22 18:08")

    with allure.step("Change sort to MORE POINTS"):
        communications.change_sort("Сначала много баллов")

    with allure.step("Check MORE POINTS in list"):
        communications.assert_call_date_and_time("09.02.22 21:23")

    with allure.step("Change sort to LESS POINTS"):
        communications.change_sort("Сначала мало баллов")

    # with allure.step("Check LESS POINTS in list"):
    #     communications.assert_call_date_and_time("10.02.22 09:51")  #10.02.22 10:57  "09.02.22 18:16" 10.02.22 09:27



@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_clear_all_fields")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("Check clear all fields by button for Ecotelecom")
def test_check_clear_all_fields(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with Ecotelecom"):
        communications.auth(ECOTELECOM, ECOPASS)

    with allure.step("Fill all default fields"):
        communications.fill_client_number("79251579005")
        communications.fill_employee_number("4995055555")
        communications.fill_client_dict_or_text("минутку", "минутку")
        communications.fill_employee_dict_or_text("Зо", "Словарь: Зомбоящик")
        communications.fill_search_length(">10")
        communications.fill_time("11:42")
        communications.fill_id("1644474236.14425")
        communications.fill_by_tag(0, "Другой отдел")

    with allure.step("Check all fields to have value"):
        expect(page.locator('[aria-label="Remove 79251579005"]')).to_be_visible()
        expect(page.locator('[aria-label="Remove 4995055555"]')).to_be_visible()
        expect(page.locator('[aria-label="Remove минутку"]')).to_be_visible()
        expect(page.locator('[aria-label="Remove Словарь: Зомбоящик"]')).to_be_visible()
        expect(page.locator('[aria-label="Remove >10"]')).to_be_visible()
        expect(page.locator('[aria-label="Remove 11:42"]')).to_be_visible()
        expect(page.locator('[aria-label="Remove 1644474236.14425"]')).to_be_visible()
        expect(page.locator('[aria-label="Remove Другой отдел"]')).to_be_visible()
        expect(page.locator(CURRENT_TEMPLATE_NAME)).to_be_visible()

    with allure.step("Press button (Clear)"):
        communications.press_clear_button()

    with allure.step("Check that cleared"):
        expect(page.locator('[aria-label="Remove 79251579005"]')).not_to_be_visible()
        expect(page.locator('[aria-label="Remove 4995055555"]')).not_to_be_visible()
        expect(page.locator('[aria-label="Remove минутку"]')).not_to_be_visible()
        expect(page.locator('[aria-label="Remove Словарь: Зомбоящик"]')).not_to_be_visible()
        expect(page.locator('[aria-label="Remove >10"]')).not_to_be_visible()
        expect(page.locator('[aria-label="Remove 11:42"]')).not_to_be_visible()
        expect(page.locator('[aria-label="Remove 1644474236.14425"]')).not_to_be_visible()
        expect(page.locator('[aria-label="Remove Другой отдел"]')).not_to_be_visible()
        expect(page.locator(CURRENT_TEMPLATE_NAME)).to_be_visible()


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_open_call_in_new_tab_by_user")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_check_open_call_in_new_tab_by_user")
def test_check_open_call_in_new_tab_by_user(base_url, page: Page, context: BrowserContext) -> None:
    communications = Communications(page)

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN, PASSWORD)
        page.wait_for_timeout(5000)

    with allure.step("Open new tab"):
        with context.expect_page() as new_tab_event:
            communications.button_share_call.click()
            new_tab=new_tab_event.value

    with allure.step("Check"):
        new_tab.wait_for_timeout(5000)
        new_tab.wait_for_selector('[alt="Imot.io loader"]', state="hidden")
        new_tab.wait_for_timeout(3000)
        expect(new_tab.locator(AUDIO_PLAYER)).to_have_count(1)
        expect(new_tab.locator('[class*="MuiAccordionSummary-content"]')).to_have_count(1)
        expect(new_tab.locator('[class*="ClientBlock_employeePhone"]')).to_have_text("0987654321")
        expect(new_tab.locator('[class*="ClientBlock_employeeDuration"]')).to_have_text("00:00:38")
        expect(new_tab.locator('[aria-label="Применение GPT правила"]')).to_have_count(1)
        expect(new_tab.locator('[aria-label="Перетегировать"]')).to_have_count(1)
        expect(new_tab.locator('[aria-label="Скачать"]')).to_have_count(1)
        expect(new_tab.locator('[aria-label="Excel экспорт"]')).to_have_count(1)
        expect(new_tab.locator('[aria-label="Скопировать публичную ссылку"]')).to_have_count(1)
        expect(new_tab.locator('[class*="styles_withAllComments_"]')).to_have_count(1)
        expect(new_tab.get_by_text("Добавить комментарий")).to_have_count(1)
        expect(new_tab.locator('[class*="_manualGroup_"]')).to_have_count(1)
        expect(new_tab.get_by_text("nlab_speech")).to_have_count(0)
        #expect(new_tab.get_by_text("Перевод")).to_have_count(1)

    with allure.step("Close context"):
        new_tab.close()

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_open_call_in_new_tab_by_admin")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_check_open_call_in_new_tab_by_admin")
def test_check_open_call_in_new_tab_by_admin(base_url, page: Page, context: BrowserContext) -> None:
    communications = Communications(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN_ADMIN, PASSWORD)
        page.wait_for_timeout(5000)

    with allure.step("Go to user"):
        communications.go_to_user(LOGIN_USER)
        page.wait_for_timeout(5000)

    with allure.step("Open new tab"):
        with context.expect_page() as new_tab_event:
            communications.button_share_call.click()
            #page.locator(BUTTON_SHARE_CALL).locator('[type="button"]').click()
            new_tab=new_tab_event.value

    with allure.step("Check"):
        new_tab.wait_for_timeout(2000)
        new_tab.wait_for_selector('[alt="Imot.io loader"]', state="hidden")
        new_tab.wait_for_timeout(2000)
        new_tab.wait_for_load_state(state="load", timeout=wait_until_visible)
        expect(new_tab.locator(AUDIO_PLAYER)).to_have_count(1)
        expect(new_tab.locator('[class*="MuiAccordionSummary-content"]')).to_have_count(1)
        expect(new_tab.locator('[class*="ClientBlock_employeePhone"]')).to_have_text("0987654321")
        expect(new_tab.locator('[class*="ClientBlock_employeeDuration"]')).to_have_text("00:00:38")
        expect(new_tab.locator('[aria-label="Применение GPT правила"]')).to_have_count(1)
        expect(new_tab.locator('[aria-label="Перетегировать"]')).to_have_count(1)
        expect(new_tab.locator('[aria-label="Скачать"]')).to_have_count(1)
        expect(new_tab.locator('[aria-label="Excel экспорт"]')).to_have_count(1)
        expect(new_tab.locator('[aria-label="Скопировать публичную ссылку"]')).to_have_count(1)
        expect(new_tab.locator('[class*="styles_withAllComments_"]')).to_have_count(1)
        expect(new_tab.get_by_text("Добавить комментарий")).to_have_count(1)
        expect(new_tab.locator('[class*="_manualGroup_"]')).to_have_count(1)
        #expect(new_tab.get_by_text("nlab_speech")).to_have_count(1, timeout=wait_until_visible)
        #expect(new_tab.get_by_text("Перевод")).to_have_count(1)


    with allure.step("Change user in new tab. https://task.imot.io/browse/DEV-3239"):
        # page.wait_for_timeout(3000)
        # communications.go_to_user("Экотелеком")
        # page.wait_for_timeout(5000)
        # page.wait_for_load_state(state="load", timeout=wait_until_visible)
        #
        # page.wait_for_selector('[value="6204e7cb599aff4f43f5d3a0"]', state="hidden")
        # page.wait_for_timeout(10000)

        new_tab.locator("#react-select-2-input").type("Экотелеком", delay=10)
        new_tab.get_by_text("Экотелеком", exact=True).click()
        new_tab.wait_for_load_state(state="load", timeout=wait_until_visible)
        new_tab.wait_for_timeout(2000)


    with allure.step("Check"):
        new_tab.wait_for_timeout(2500)
        expect(new_tab.locator(AUDIO_PLAYER)).to_have_count(1)
        expect(new_tab.locator('[class*="MuiAccordionSummary-content"]')).to_have_count(1)
        expect(new_tab.locator('[class*="ClientBlock_employeePhone"]')).to_have_text("0987654321")
        expect(new_tab.locator('[class*="ClientBlock_employeeDuration"]')).to_have_text("00:00:38")
        expect(new_tab.locator('[aria-label="Применение GPT правила"]')).to_have_count(1)
        expect(new_tab.locator('[aria-label="Перетегировать"]')).to_have_count(1)
        expect(new_tab.locator('[aria-label="Скачать"]')).to_have_count(1)
        expect(new_tab.locator('[aria-label="Excel экспорт"]')).to_have_count(1)
        expect(new_tab.locator('[aria-label="Скопировать публичную ссылку"]')).to_have_count(1)
        expect(new_tab.locator('[class*="styles_withAllComments_"]')).to_have_count(1)
        expect(new_tab.get_by_text("Добавить комментарий")).to_have_count(1)
        expect(new_tab.locator('[class*="_manualGroup_"]')).to_have_count(1)

    with allure.step("Close context"):
        new_tab.close()

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_content_button_calls_actions_for_user")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_check_content_button_calls_actions (...) for user")
def test_check_content_button_calls_actions_for_user(base_url, page: Page) -> None:
    communications = Communications(page)

    options_list = ("Применить GPTПоменять аудио каналыЗагрузить теги из crmПрименить информированиеПрименить адресную"
                    " книгуФильтр теговПеревыгрузить из интеграцииОбработать непрерывные записиВосстановить непрерывные записи")

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN, PASSWORD)

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_less_than_50()

    with allure.step("Press button (Calls action)"):
        communications.press_calls_action_button_in_list(0)

    with allure.step("Check content of button (...) calls action"):
        expect(communications.menu).to_have_text(options_list)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_download_button_in_calls_list")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_check_content_button_calls (download)")
def test_check_download_button_in_calls_list(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with Ecotelecom"):
        communications.auth(ECOTELECOM, ECOPASS)

    with allure.step("Choose period from 01/01/2022 to 31/12/2022"):
        communications.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Fill ID to find call"):
        communications.fill_id("1644268426.90181")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_less_than_50()

    with allure.step("Press button (Download)"):
        communications.press_calls_list_download_button(0)

    with allure.step("Check content of button download"):
        expect(communications.menu).to_have_text("Экспорт аудиоЭкспорт аудио (многоканальное)Экспорт расшифровкиЭкспорт коммуникаций")

    with allure.step("Choose (Export audio) option from opened menu"):
        # Start waiting for the download
        with page.expect_download(timeout=60000) as download_info:
            # Perform the action that initiates download
            communications.menu.get_by_text("Экспорт аудио", exact=True).click()
        download = download_info.value
        path = f'{os.getcwd()}/'

        # Wait for the download process to complete and save the downloaded file somewhere
        download.save_as(path + download.suggested_filename)

    with allure.step("Check that export (zip) downloaded"):
        assert os.path.isfile(path + download.suggested_filename) == True
        assert 120000 < os.path.getsize(path + download.suggested_filename) < 160000

    with allure.step("Remove downloaded export (zip)"):
        os.remove(path + download.suggested_filename)

    with allure.step("Check that downloaded export (zip) removed"):
        assert os.path.isfile(path + download.suggested_filename) == False

    with allure.step("Press button (Download)"):
        communications.press_calls_list_download_button(0)

    with allure.step("Choose (Export audio) option from opened menu"):
        # Start waiting for the download
        with page.expect_download(timeout=60000) as download_info:
            # Perform the action that initiates download
            communications.menu.get_by_text("Экспорт аудио (многоканальное)", exact=True).click()
        download = download_info.value
        path = f'{os.getcwd()}/'

        # Wait for the download process to complete and save the downloaded file somewhere
        download.save_as(path + download.suggested_filename)

    with allure.step("Check that export (zip) downloaded"):
        assert os.path.isfile(path + download.suggested_filename) == True
        assert 500000 < os.path.getsize(path + download.suggested_filename) < 600000

    with allure.step("Remove downloaded export (zip)"):
        os.remove(path + download.suggested_filename)

    with allure.step("Check that downloaded export (zip) removed"):
        assert os.path.isfile(path + download.suggested_filename) == False

    with allure.step("Press button (Download)"):
        communications.press_calls_list_download_button(0)

    with allure.step("Press (Export)"):
        communications.menu.get_by_text("Экспорт расшифровки", exact=True).click()
        page.wait_for_selector(MODAL_WINDOW)

    with allure.step("Choose (Export transcribe) option from opened menu"):
        # Start waiting for the download
        with page.expect_download(timeout=60000) as download_info:
            # Perform the action that initiates download
            communications.modal_window.get_by_text("Экспортировать", exact=True).click()
        download = download_info.value
        path = f'{os.getcwd()}/'
        # Wait for the download process to complete and save the downloaded file somewhere
        download.save_as(path + download.suggested_filename)

    with allure.step("Check that export (zip) downloaded"):
        assert os.path.isfile(path + download.suggested_filename) == True
        assert 8000 < os.path.getsize(path + download.suggested_filename) < 20000

    with allure.step("Check what we have inside excel"):
        wb = load_workbook(path + download.suggested_filename)
        sheet = wb.active

        assert sheet["A1"].value == "Экспорт расшифровки звонков"
        assert sheet.max_row == 55
        assert sheet.max_column == 4

    with allure.step("Remove downloaded export (zip)"):
        os.remove(path + download.suggested_filename)

    with allure.step("Check that downloaded export (zip) removed"):
        assert os.path.isfile(path + download.suggested_filename) == False

    with allure.step("Close modal with export"):
        communications.close_export_modal()

    with allure.step("Press button (Download)"):
        communications.press_calls_list_download_button(0)

    with allure.step("Press (Export)"):
        communications.menu.get_by_text("Экспорт коммуникаций", exact=True).click()
        page.wait_for_selector(MODAL_WINDOW)

    with allure.step("Choose (Export transcribe) option from opened menu"):
        # Start waiting for the download
        with page.expect_download(timeout=60000) as download_info:
            # Perform the action that initiates download
            communications.modal_window.get_by_text("Экспортировать", exact=True).click()
        download = download_info.value
        path = f'{os.getcwd()}/'

        # Wait for the download process to complete and save the downloaded file somewhere
        download.save_as(path + download.suggested_filename)

    with allure.step("Check that export (zip) downloaded"):
        assert os.path.isfile(path + download.suggested_filename) == True
        assert 6000 < os.path.getsize(path + download.suggested_filename) < 7000

    with allure.step("Check what we have inside excel"):
        wb = load_workbook(path + download.suggested_filename)
        sheet = wb.active

        assert sheet["A1"].value == "Выгрузка списка звонков"
        assert sheet.max_row == 3
        assert sheet.max_column == 35

    with allure.step("Remove downloaded export (zip)"):
        os.remove(path + download.suggested_filename)

    with allure.step("Check that downloaded export (zip) removed"):
        assert os.path.isfile(path + download.suggested_filename) == False

    with allure.step("Close modal with export"):
        communications.close_export_modal()



@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_buttons_in_open_call")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_check_buttons_in_open_call")
def test_check_buttons_in_open_call(base_url, page: Page) -> None:
    communications = Communications(page)

    text = ("Поменять аудио каналыПрименить адресную книгуПрименить информированиеУдаленные тегиЗагрузить теги из "
            "crmПеревыгрузить из интеграцииПоказать скрытые тегиМета инфоРедактировать правило оповещения")

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN, PASSWORD)

    # with allure.step("Press button (Find communications)"):
    #     press_find_communications(page)

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Check that all 6 buttons in expanded call visible"):
        expect(page.locator(OPEN_CALL_AREA).locator('[aria-label="Переход в источник коммуникации"]')).to_be_visible()
        expect(page.locator(OPEN_CALL_AREA).locator('[aria-label="Применение GPT правила"]')).to_be_visible()
        expect(page.locator(OPEN_CALL_AREA).locator('[aria-label="Перетегировать"]')).to_be_visible()
        expect(page.locator(OPEN_CALL_AREA).locator('[aria-label="Скачать"]')).to_be_visible()
        expect(page.locator(OPEN_CALL_AREA).locator('[aria-label="Excel экспорт"]')).to_be_visible()
        expect(page.locator(OPEN_CALL_AREA).locator('[aria-label="Скопировать публичную ссылку"]')).to_be_visible()
        expect(page.locator(OPEN_CALL_AREA).locator(BUTTON_CALLS_ACTION)).to_be_visible()

    with allure.step("Click button (calls action)"):
        page.locator(OPEN_CALL_AREA).locator(BUTTON_CALLS_ACTION).locator('[type="button"]').click()

    with (allure.step("Check content in opened menu")):
        expect(page.locator(OPEN_CALL_AREA).locator(MENU)).to_have_text(text)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)

@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_download_call_from_expanded_call")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_check_download_call_from_expanded_call")
def test_check_download_call_from_expanded_call(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN, PASSWORD)

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Press (Download) button and download file"):
        # Start waiting for the download
        with page.expect_download(timeout=60000) as download_info:
            # Perform the action that initiates download
            page.wait_for_timeout(1000)
            page.locator(OPEN_CALL_AREA).locator('[aria-label="Скачать"]').locator('[type="button"]').click()
        download = download_info.value
        path = f'{os.getcwd()}/'

        # Wait for the download process to complete and save the downloaded file somewhere
        download.save_as(path + download.suggested_filename)

    with allure.step("Check that file opus downloaded"):
        assert os.path.isfile(path + download.suggested_filename) == True
        assert 140000 < os.path.getsize(path + download.suggested_filename) < 148000

    with allure.step("Remove downloaded file"):
        os.remove(path + download.suggested_filename)

    with allure.step("Check that file removed"):
        assert os.path.isfile(path + download.suggested_filename) == False

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)

@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_download_excel_from_expanded_call")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_check_download_excel_from_expanded_call")
def test_check_download_excel_from_expanded_call(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with Ecotelecom"):
        communications.auth(ECOTELECOM, ECOPASS)

    with allure.step("Choose period from 01/01/2022 to 31/12/2022"):
        communications.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Fill first ID to find call"):
        communications.fill_id("1644268426.90181")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_less_than_50()

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Press (EX) button"):
        communications.press_ex_button_in_expanded_call()

    with allure.step("Check content of modal window"):
        expect(communications.modal_window.get_by_text("Теги без значений")).to_have_count(1)
        expect(communications.modal_window.get_by_text("Теги со значениями")).to_have_count(1)
        expect(communications.modal_window.get_by_text("Параметры коммуникаций")).to_have_count(1)

    with allure.step(" and download excel"):
        # Start waiting for the download
        with page.expect_download(timeout=60000) as download_info:
            # Perform the action that initiates download
            communications.modal_window.locator('[class*="buttonsBlock_"]').get_by_role("button", name="Экспортировать").click()
        download = download_info.value
        path = f'{os.getcwd()}/'

        # Wait for the download process to complete and save the downloaded file somewhere
        download.save_as(path + download.suggested_filename)

    with allure.step("Check that excel export downloaded"):
        assert os.path.isfile(path + download.suggested_filename) == True
        assert os.path.getsize(path + download.suggested_filename) > 7300

    with allure.step("Check what we have inside excel"):
        wb = load_workbook(path + download.suggested_filename)
        sheet = wb.active

        assert sheet["A1"].value == "Экспорт расшифровки звонков"
        assert sheet.max_row == 55
        assert sheet.max_column == 4

    with allure.step("Remove downloaded excel export"):
        os.remove(path + download.suggested_filename)

    with allure.step("Check that excel export removed"):
        assert os.path.isfile(path + download.suggested_filename) == False

    with allure.step("Close export modal"):
        communications.close_export_modal()

    with allure.step("Fill second ID to find call"):
        page.wait_for_selector(INPUT_ID, timeout=wait_until_visible)
        page.locator(INPUT_ID).locator('[type="text"]').clear()
        page.locator(INPUT_ID).locator('[type="text"]').type("1644268692.90190", delay=10)
        page.wait_for_timeout(500)

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_less_than_50()

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Press (EX) button"):
        communications.press_ex_button_in_expanded_call()

    with allure.step("Check content of modal window"):
        expect(communications.modal_window.get_by_text("Теги без значений")).to_have_count(1)
        expect(communications.modal_window.get_by_text("Теги со значениями")).to_have_count(1)
        expect(communications.modal_window.get_by_text("Параметры коммуникаций")).to_have_count(1)

    with allure.step(" and download excel"):
        # Start waiting for the download
        with page.expect_download(timeout=60000) as download_info:
            # Perform the action that initiates download
            communications.modal_window.locator('[class*="buttonsBlock_"]').get_by_role("button", name="Экспортировать").click()
        download = download_info.value
        path = f'{os.getcwd()}/'

        # Wait for the download process to complete and save the downloaded file somewhere
        download.save_as(path + download.suggested_filename)

    with allure.step("Check that excel export downloaded"):
        assert os.path.isfile(path + download.suggested_filename) == True
        assert os.path.getsize(path + download.suggested_filename) > 7100

    with allure.step("Check what we have inside excel"):
        wb = load_workbook(path + download.suggested_filename)
        sheet = wb.active

        assert sheet["A1"].value == "Экспорт расшифровки звонков"
        assert sheet.max_row == 47
        assert sheet.max_column == 4

    with allure.step("Remove downloaded excel export"):
        os.remove(path + download.suggested_filename)

    with allure.step("Check that excel export removed"):
        assert os.path.isfile(path + download.suggested_filename) == False



@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_search_template")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_check_search_template")
def test_check_search_template(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN, PASSWORD)

    with allure.step("Check that no any templates"):
        communications.assert_template_name("Сохраненные шаблоны поиска(0)")

    with allure.step("Save template"):
        communications.press_save_template()

    with allure.step("Check that (add) button disabled"):
        expect(page.locator(MODAL_WINDOW).locator(BUTTON_SUBMIT)).to_be_disabled()

    with allure.step("Fill template name"):
        communications.fill_name("firstTemplate")

    with allure.step("Check that (add) button enabled"):
        expect(page.locator(MODAL_WINDOW).locator(BUTTON_SUBMIT)).to_be_enabled()

    with allure.step("Press (add)"):
        page.locator(MODAL_WINDOW).locator(BUTTON_SUBMIT).click()

    with allure.step("Check that template saved"):
        communications.assert_template_name("firstTemplate(1)")

    with allure.step("Rename template"):
        communications.press_rename_template()

    with allure.step("Check that (add) button disabled"):
        expect(page.locator(MODAL_WINDOW).locator(BUTTON_SUBMIT)).to_be_disabled()

    with allure.step("Fill template name"):
        communications.fill_name("renameTemplate")

    with allure.step("Check that (add) button enabled"):
        expect(page.locator(MODAL_WINDOW).locator(BUTTON_SUBMIT)).to_be_enabled()

    with allure.step("Press (add)"):
        page.locator(MODAL_WINDOW).locator(BUTTON_SUBMIT).click()

    with allure.step("Check that template saved"):
        communications.assert_template_name("renameTemplate(1)")

    with allure.step("Delete template"):
        communications.press_delete_template()

    with allure.step("Press (cancel)"):
        page.get_by_role("button", name="Отмена").click()

    with allure.step("Delete template"):
        communications.press_delete_template()

    with allure.step("Confirm delete"):
        page.locator(MODAL_WINDOW).locator(BUTTON_SUBMIT).click()
        page.wait_for_timeout(2000)

    with allure.step("Check that template saved"):
        communications.assert_template_name("Сохраненные шаблоны поиска(0)")

    # with allure.step("Check stupid text"):
    #     expect(page.locator('[style="font-size: 13px; margin-top: 15px;"]').get_by_text("Поиск пуст. Добавить фильтры можно с помощь 'Изменить фильтры'")).to_be_visible()

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_communication_comment")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_check_communication_comment")
def test_check_communication_comment(base_url, page: Page) -> None:
    communications = Communications(page)

    today = datetime.now().strftime("%d.%m.%Y, ")  # %H:%M can fail test if minutes changed while test running

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN, PASSWORD)

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Press (add comment)"):
        communications.press_add_comment()

    with allure.step("Check that comment form openned"):
        page.wait_for_timeout(1000)
        expect(page.locator(ALL_COMMENTS_AREA).locator('[type="button"]').nth(1)).to_be_disabled()
        expect(page.locator(ALL_COMMENTS_AREA).locator(CHECKBOX)).not_to_be_checked()

    with allure.step("Check that we can close comment form with X"):
        page.locator(BUTTON_CLOSE).click()

    with allure.step("Press (add comment)"):
        communications.press_add_comment()

    with allure.step("Check checkbox (hidden)"):
        page.locator(ALL_COMMENTS_AREA).locator(CHECKBOX).check()

    with allure.step("Add title"):
        page.locator(BUTTON_ADD_COMMENT_TITLE).click()

    with allure.step("Fill title"):
        page.locator('[id*="post_comment_"]').locator('[class*="styles_title_"]').type("CommentTitle", delay=10)

    with allure.step("Check that button (add comment) still disabled"):
        expect(page.locator(ALL_COMMENTS_AREA).locator('[type="button"]').nth(1)).to_be_disabled()

    with allure.step("Fill comment"):
        page.locator('[class*="styles_textareaWrapper"]').locator('[class*="styles_textarea_"]').type("CommentText", delay=10)

    with allure.step("Press (add comment)"):
        page.locator(ALL_COMMENTS_AREA).locator('[type="button"]').nth(1).click()
        page.wait_for_timeout(1000)

    with allure.step("Check that comment saved and saved right"):
        expect(page.locator('[class*="styles_author_"]')).to_have_text(LOGIN)
        expect(page.locator('[class*="styles_time_"]')).to_contain_text(today)
        expect(page.locator(COMMENT_AREA).locator('[class*="styles_head_"]').locator('[height="18"]')).to_have_count(1)
        expect(page.locator(COMMENT_AREA).locator('[class*="styles_title_"]')).to_have_text("CommentTitle")
        expect(page.locator(COMMENT_AREA).locator('[class*="styles_message_"]')).to_have_text("CommentText")
        expect(page.locator(COMMENT_AREA).locator('[class*="styles_whose"]')).to_have_text("Коммуникация")

    with allure.step("Press to (...) comment options"):
        page.locator('[class*="styles_optionsSelect_"]').click()

    with allure.step("Choose and click (edit)"):
        communications.menu.get_by_text("Редактировать", exact=True).click()
        page.wait_for_selector('[class*="styles_checkButton_"]')
        expect(page.locator('[class*="styles_checkButton_"]')).to_be_disabled()

    with allure.step("Edit title"):
        page.locator('[class*="styles_editableInput"]').clear()
        page.locator('[class*="styles_editableInput"]').type("EditedTitle", delay=10)

    with allure.step("Edit comment text"):
        page.locator('[class*="styles_editableTextarea"]').clear()
        page.locator('[class*="styles_editableTextarea"]').type("EditedText", delay=10)

    with allure.step("Save edited comment"):
        page.locator('[class*="styles_checkButton_"]').click()

    with allure.step("Check that edition comment saved"):
        expect(page.locator(COMMENT_AREA).locator('[class*="styles_title_"]')).to_have_text("EditedTitle")
        expect(page.locator(COMMENT_AREA).locator('[class*="styles_message_"]')).to_have_text("EditedText")

    with allure.step("Press to (...) comment options"):
        page.locator('[class*="styles_optionsSelect_"]').click()
        page.wait_for_selector(MENU)

    with allure.step("Choose and click (delete)"):
        communications.choose_from_menu_by_text_and_wait_for_modal("Удалить комментарий")

    with allure.step("Confirm deleting"):
        communications.modal_window.get_by_role("button", name="Удалить").click()
        page.wait_for_selector(MODAL_WINDOW, state="hidden")
        page.wait_for_timeout(500)

    with allure.step("Check that comment was deleted"):
        expect(page.locator(COMMENT_AREA)).not_to_be_visible()

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_re_recognize_for_call_list")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check re-recognize in call list for ecotelecom. First finding by ID call with 0 length")
def test_check_re_recognize_for_call_list(base_url, page: Page) -> None:
    communications = Communications(page)

    # expected_languages = ("Английский (Великобритания)Английский (США)Испанский (Латинская Америка, Карибский регион, "
    #                       "код региона UN M49)Испанский (Испания)Французский (Франция)Португальский "
    #                       "(Бразилия)Португальский (Португалия)РусскийТурецкий (Турция)УкраинскийУзбекскийАвто")

    #expected_engines = "DeepgramGigaAMHappyscribeNLab SpeechIMOT.IOwhisperЯндексЯндекс v3"
    #expected_engines = "DeepgramgigaamHappyscribenexaraNLab SpeechIMOT.IOwhisperЯндексyandex_v3"


    alert_merge = "Опция 'Объединить дорожки в один файл' не может быть выбрана одновременно с любой из диаризаций"

    alert_diarization = "Выберите только одну диаризацию среди опций"

    action_started = "Действие начато"

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Auth with admin"):
        communications.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        communications.go_to_user(LOGIN_USER)

    with allure.step("Click calls list actions button (...)"):
        communications.press_calls_action_button_in_list(0)

    with allure.step("Choose re-recognize in menu"):
        communications.choose_from_menu_by_text_and_wait_for_modal("Перераспознать")

    with allure.step("Check modal window content"):
        expect(page.locator('[class*="styles_sttAllFoudCalls_"]')).to_contain_text(" (количество коммуникаций:  1)")
        expect(page.locator(SELECT_LANGUAGE)).to_contain_text("Русский")
        expect(page.locator(MODAL_WINDOW).locator(BUTTON_CROSS)).to_have_count(1)
        expect(page.locator(MODAL_WINDOW).locator(BUTTON_SUBMIT)).to_contain_text("Перераспознать")

#  check all combinations of engines and models

    with allure.step("Click to language"):
        communications.click_language_select()

    with allure.step("Check language list"):
        communications.assert_menu_values(fucking_stupidity)

    with allure.step("Close language menu"):
        page.locator('[class*="STT_order_"]').click()

    # for ecotelecom engine and model already selected
    # with allure.step("Check that engine not selected"):
    #     expect(page.locator(SELECT_ENGINE)).to_contain_text("Выберите движок")

    # with allure.step("Check that model not selected"):
    #     expect(page.locator(SELECT_MODEL)).to_contain_text("Выберите модель")

    with allure.step("Click to engine"):
        communications.click_engine_select()

    with allure.step("Check engine list"):
        communications.assert_menu_values(expected_engines)

    # with allure.step("Click to engine"):
    #     communications.click_engine_select()

    with allure.step("Choose Deepgram"):
        communications.choose_option(1)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Check model list"):
        communications.assert_menu_values("Обобщённаяwhisper")

    with allure.step("Select model Обобщённая"):
        communications.choose_option(0)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Select model whisper"):
        communications.choose_option(1)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("Deepgram")
        expect(page.locator(SELECT_MODEL)).to_contain_text("whisper")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator(CHECKBOX)).to_have_count(3)

    with allure.step("Click to engine"):
        communications.click_engine_select()

    with allure.step("Choose Happyscribe"):
        communications.choose_option(3)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Check model list"):
        communications.assert_menu_values("Стандарт")

    with allure.step("Select model Стандарт"):
        communications.choose_option(0)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("Happyscribe")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Стандарт")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator(CHECKBOX)).to_have_count(3)

    with allure.step("Click to engine"):
        communications.click_engine_select()

    with allure.step("Choose NLab Speech"):
        communications.choose_option(4)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Check model list"):
        communications.assert_menu_values("ОбобщённаяЖадный")

    with allure.step("Select model Обобщённая"):
        communications.choose_option(0)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Select model Жадный"):
        communications.choose_option(1)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("NLab Speech")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Жадный")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator(CHECKBOX)).to_have_count(3)

    with allure.step("Click to engine"):
        communications.click_engine_select()

    with allure.step("Choose IMOT.IO"):
        communications.choose_option(5)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Check model list"):
        communications.assert_menu_values("Стандарт")

    with allure.step("Select model Стандарт"):
        communications.choose_option(0)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("IMOT.IO")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Стандарт")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator(CHECKBOX)).to_have_count(3)

    with allure.step("Click to engine"):
        communications.click_engine_select()

    with allure.step("Choose whisper"):
        communications.choose_option(6)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Check model list"):
        communications.assert_menu_values("Стандарт")

    with allure.step("Select model Стандарт"):
        communications.choose_option(0)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("whisper")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Стандарт")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator(CHECKBOX)).to_have_count(5)

    with allure.step("Click to engine"):
        communications.click_engine_select()

    with allure.step("Choose Яндекс"):
        communications.choose_option(7)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Check model list"):
        communications.assert_menu_values("Отложенная обобщённаяОбобщённая")

    with allure.step("Select model Отложенная обобщённая"):
        communications.choose_option(0)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Select model Обобщённая"):
        communications.choose_option(1)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("Яндекс")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Обобщённая")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator(CHECKBOX)).to_have_count(3)

    with allure.step("Click to engine"):
        communications.click_engine_select()

    with allure.step("Choose yandex_v3"):
        communications.choose_option(8)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Check model list"):
        communications.assert_menu_values("Отложенная обобщённаяОбобщённая")

    with allure.step("Select model Отложенная обобщённая"):
        communications.choose_option(0)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Select model Обобщённая"):
        communications.choose_option(1)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("Яндекс v3")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Обобщённая")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ENGINE_DIARIZATION)).to_be_checked()
        expect(page.locator(CHECKBOX_NORMALIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_PROFANITY_FILTER)).not_to_be_checked()
        expect(page.locator(CHECKBOX_LITERATURE_STYLE)).not_to_be_checked()
        expect(page.locator(CHECKBOX_PHONE_FORMATTING)).to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator(CHECKBOX)).to_have_count(8)

#  check save combinations

    with allure.step("Click to engine"):
        communications.click_engine_select()

    with allure.step("Choose yandex_v3"):
        communications.choose_option(8)

    with allure.step("Check (Save) button is disabled"):
        expect(page.locator(BLOCK_WITH_BUTTON).locator(BUTTON_SUBMIT)).to_be_enabled()

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Select model best"):
        communications.choose_option(0)

    with allure.step("Check (Save) button is enabled"):
        expect(page.locator(BLOCK_WITH_BUTTON).locator(BUTTON_SUBMIT)).to_be_enabled()

    with allure.step("Check merge all to one checkbox"):
        page.locator(CHECKBOX_MERGE_ALL_TO_ONE).set_checked(checked=True)

    with allure.step("Try to save"):
        communications.click_submit_in_word_processing()

    with allure.step("Wait for alert and check alert message"):
        communications.check_alert(alert_merge)

    with allure.step("Uncheck merge all to one checkbox"):
        page.locator(CHECKBOX_MERGE_ALL_TO_ONE).uncheck()

    with allure.step("Check diarization checkbox"):
        page.locator(CHECKBOX_DIARIZATION).set_checked(checked=True)

    with allure.step("Try to save"):
        communications.click_submit_in_word_processing()

    with allure.step("Wait for alert and check alert message"):
        communications.check_alert(alert_diarization)

    with allure.step("Uncheck diarization checkbox"):
        page.locator(CHECKBOX_DIARIZATION).uncheck()

    with allure.step("Change parameters"):
        page.locator(RECOGNITION_PRIORITY).locator('[type="number"]').fill("10")
        page.locator(CHECKBOX_ECONOMIZE).set_checked(checked=True)
        #page.locator(CHECKBOX_USE_WEBHOOK).set_checked(checked=True)

    with allure.step("Press (Save)"):
        communications.click_submit_in_word_processing()

    with allure.step("Wait for alert and check alert message"):
        page.wait_for_selector('[class*="SnackbarItem"]', timeout=wait_until_visible)
        # expect(page.locator('[class*="SnackbarItem"]')).to_contain_text(action_started)
        # page.wait_for_selector('[class*="SnackbarItem"]', state="hidden", timeout=wait_until_visible)

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)

@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_re_recognize_for_expanded_call")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check re-recognize in expanded call for ecotelecom. First finding by ID call with 0 length")
def test_check_re_recognize_for_expanded_call(base_url, page: Page) -> None:
    communications = Communications(page)

    # expected_languages = ("Английский (Великобритания)Английский (США)Испанский (Латинская Америка, Карибский регион, "
    #                       "код региона UN M49)Испанский (Испания)Французский (Франция)Португальский "
    #                       "(Бразилия)Португальский (Португалия)РусскийТурецкий (Турция)УкраинскийУзбекскийАвто")

    # expected_engines = "DeepgramGigaAMHappyscribeNLab SpeechIMOT.IOwhisperЯндексЯндекс v3"
    # expected_engines = "DeepgramgigaamHappyscribenexaraNLab SpeechIMOT.IOwhisperЯндексyandex_v3"

    alert_merge = "Опция 'Объединить дорожки в один файл' не может быть выбрана одновременно с любой из диаризаций"

    alert_diarization = "Выберите только одну диаризацию среди опций"

    action_started = "Действие начато"

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with admin"):
        communications.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        communications.go_to_user(LOGIN_USER)

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Click calls list actions button (...)"):
        communications.press_calls_action_button_in_list(1)

    with allure.step("Choose re-recognize in menu"):
        communications.choose_from_menu_by_text_and_wait_for_modal("Перераспознать")

    with allure.step("Check modal window content"):
        expect(page.locator(SELECT_LANGUAGE)).to_contain_text("Русский")
        expect(page.locator(MODAL_WINDOW).locator(BUTTON_CROSS)).to_have_count(1)
        expect(page.locator(MODAL_WINDOW).locator(BUTTON_SUBMIT)).to_contain_text("Перераспознать")
#  check all combinations of engines and models

    with allure.step("Click to language"):
        communications.click_language_select()

    with allure.step("Check language list"):
        communications.assert_menu_values(fucking_stupidity)

    with allure.step("Close language menu"):
        page.locator('[class*="STT_order_"]').click()

    # for ecotelecom engine and model already selected
    # with allure.step("Check that engine not selected"):
    #     expect(page.locator(SELECT_ENGINE)).to_contain_text("Выберите движок")

    # with allure.step("Check that model not selected"):
    #     expect(page.locator(SELECT_MODEL)).to_contain_text("Выберите модель")

    with allure.step("Click to engine"):
        communications.click_engine_select()

    with allure.step("Check engine list"):
        communications.assert_menu_values(expected_engines)

    # with allure.step("Click to engine"):
    #     communications.click_engine_select()

    with allure.step("Choose Deepgram"):
        communications.choose_option(1)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Check model list"):
        communications.assert_menu_values("Обобщённаяwhisper")

    with allure.step("Select model Обобщённая"):
        communications.choose_option(0)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Select model whisper"):
        communications.choose_option(1)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("Deepgram")
        expect(page.locator(SELECT_MODEL)).to_contain_text("whisper")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator(CHECKBOX)).to_have_count(3)

    with allure.step("Click to engine"):
        communications.click_engine_select()

    with allure.step("Choose Happyscribe"):
        communications.choose_option(3)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Check model list"):
        communications.assert_menu_values("Стандарт")

    with allure.step("Select model Стандарт"):
        communications.choose_option(0)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("Happyscribe")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Стандарт")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator(CHECKBOX)).to_have_count(3)

    with allure.step("Click to engine"):
        communications.click_engine_select()

    with allure.step("Choose NLab Speech"):
        communications.choose_option(4)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Check model list"):
        communications.assert_menu_values("ОбобщённаяЖадный")

    with allure.step("Select model Обобщённая"):
        communications.choose_option(0)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Select model Жадный"):
        communications.choose_option(1)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("NLab Speech")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Жадный")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator(CHECKBOX)).to_have_count(3)
    #
    with allure.step("Click to engine"):
        communications.click_engine_select()

    with allure.step("Choose IMOT.IO"):
        communications.choose_option(5)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Check model list"):
        communications.assert_menu_values("Стандарт")

    with allure.step("Select model Стандарт"):
        communications.choose_option(0)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("IMOT.IO")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Стандарт")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator(CHECKBOX)).to_have_count(3)

    with allure.step("Click to engine"):
        communications.click_engine_select()

    with allure.step("Choose whisper"):
        communications.choose_option(6)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Check model list"):
        communications.assert_menu_values("Стандарт")

    with allure.step("Select model Стандарт"):
        communications.choose_option(0)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("whisper")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Стандарт")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator(CHECKBOX)).to_have_count(5)

    with allure.step("Click to engine"):
        communications.click_engine_select()

    with allure.step("Choose Яндекс"):
        communications.choose_option(7)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Check model list"):
        communications.assert_menu_values("Отложенная обобщённаяОбобщённая")

    with allure.step("Select model Отложенная обобщённая"):
        communications.choose_option(0)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Select model Обобщённая"):
        communications.choose_option(1)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("Яндекс")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Обобщённая")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator(CHECKBOX)).to_have_count(3)
    #
    with allure.step("Click to engine"):
        communications.click_engine_select()

    with allure.step("Choose yandex_v3"):
        communications.choose_option(8)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Check model list"):
        communications.assert_menu_values("Отложенная обобщённаяОбобщённая")

    with allure.step("Select model Отложенная обобщённая"):
        communications.choose_option(0)

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Select model Обобщённая"):
        communications.choose_option(1)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("Яндекс v3")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Обобщённая")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ENGINE_DIARIZATION)).to_be_checked()
        expect(page.locator(CHECKBOX_NORMALIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_PROFANITY_FILTER)).not_to_be_checked()
        expect(page.locator(CHECKBOX_LITERATURE_STYLE)).not_to_be_checked()
        expect(page.locator(CHECKBOX_PHONE_FORMATTING)).to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator(CHECKBOX)).to_have_count(8)

#  check save combinations

    with allure.step("Click to engine"):
        communications.click_engine_select()

    with allure.step("Choose assembly_ai"):
        communications.choose_option(8)

    with allure.step("Check (Save) button is disabled"):
        expect(page.locator(BLOCK_WITH_BUTTON).locator(BUTTON_SUBMIT)).to_be_enabled()

    with allure.step("Click to model"):
        communications.click_model_select()

    with allure.step("Select model best"):
        communications.choose_option(0)

    with allure.step("Check (Save) button is enabled"):
        expect(page.locator(BLOCK_WITH_BUTTON).locator(BUTTON_SUBMIT)).to_be_enabled()

    with allure.step("Check merge all to one checkbox"):
        page.locator(CHECKBOX_MERGE_ALL_TO_ONE).set_checked(checked=True)

    with allure.step("Try to save"):
        communications.click_submit_in_word_processing()

    with allure.step("Wait for alert and check alert message"):
        communications.check_alert(alert_merge)

    with allure.step("Uncheck merge all to one checkbox"):
        page.locator(CHECKBOX_MERGE_ALL_TO_ONE).uncheck()

    with allure.step("Check diarization checkbox"):
        page.locator(CHECKBOX_DIARIZATION).set_checked(checked=True)

    with allure.step("Try to save"):
        communications.click_submit_in_word_processing()

    with allure.step("Wait for alert and check alert message"):
        communications.check_alert(alert_diarization)

    with allure.step("Uncheck diarization checkbox"):
        page.locator(CHECKBOX_DIARIZATION).uncheck()

    with allure.step("Change parameters"):
        page.locator(RECOGNITION_PRIORITY).locator('[type="number"]').fill("10")
        page.locator(CHECKBOX_ECONOMIZE).set_checked(checked=True)
        #page.locator(CHECKBOX_USE_WEBHOOK).set_checked(checked=True)

    with allure.step("Press (Save)"):
        communications.click_submit_in_word_processing()

    with allure.step("Wait for alert and check alert message"):
        communications.check_alert(action_started)

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_communication_manual_tag")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_check_communication_manual_tag")
def test_check_communication_manual_tag(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN, PASSWORD)

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Press (add manual tag)"):
        communications.press_cross_in_manual_tags()

    with allure.step("Press (Enter) with empty input"):
        communications.press_key("Enter")

    with allure.step("Wait for alert and check alert message"):
        communications.check_alert("Укажите название тега")

    # with allure.step("Press (add manual tag)"):
    #     communications.press_cross_in_manual_tags()

    with allure.step("Press (New tag)"):
        page.wait_for_timeout(1000)
        page.locator(SELECT_WITH_SEARCH_MANUAL_TAGS).locator('[class*="_tagGhost_"]').click()
        page.wait_for_timeout(1000)
        page.wait_for_selector(SELECT_WITH_SEARCH_MANUAL_TAGS)

    with allure.step("Add manual tag name"):
        communications.add_manual_tag_name("manual_tag")

    with allure.step("Press (add comment)"):
        communications.press_key("Enter")

    with allure.step("Wait for alert and check alert message"):
        communications.check_alert("Тег успешно добавлен")

    with allure.step("Check that we can see tags"):
        expect(page.get_by_text("manual_tag")).to_have_count(2)
        #communications.assert_tags_have_count(4, 1)

    with allure.step("Delete manual tag from call header"):
        communications.delete_manual_tag_from_call_header(0)

    with allure.step("Wait for alert and check alert message"):
        communications.check_alert("Тег удален")

    with allure.step("Check that we can see tags"):
        expect(page.get_by_text("manual_tag")).to_have_count(0)
        #communications.assert_tags_have_count(2, 1)
    #

    # with allure.step("Kostyl for https://task.imot.io/browse/DEV-3083"):
    #     page.locator('[class*="_manualGroup_"]').locator('[type="button"]').click()
    #     page.wait_for_timeout(500)

    with allure.step("Press (add manual tag)"):
        communications.press_cross_in_manual_tags()

    with allure.step("Add manual tag name"):
        communications.add_manual_tag_name("manual_tag")

    with allure.step("Press (add comment)"):
        communications.press_key("Enter")

    with allure.step("Wait for alert and check alert message"):
        communications.check_alert("Тег успешно добавлен")

    with allure.step("Check that we can see tags"):
        expect(page.get_by_text("manual_tag")).to_have_count(2)
        #communications.assert_tags_have_count(4, 1)

    with allure.step("Delete manual tag from manual tags"):
        communications.delete_manual_tag_from_manual_tags(0)

    with allure.step("Wait for alert and check alert message"):
        communications.check_alert("Тег удален")

    with allure.step("Check that we can see tags"):
        expect(page.get_by_text("manual_tag")).to_have_count(0)
        #communications.assert_tags_have_count(2, 1)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_search_and_switch_to_other_user")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Searching all communications for Ecotelecom and switch to auto_test_user and check that calls changed")
def test_check_search_and_switch_to_other_user(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with admin"):
        communications.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Change user to auto_test_user"):
        communications.go_to_user(LOGIN_USER)

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_less_than_50()

    with allure.step("Check that 1 communications found. Should be 1 and date today"):
        communications.assert_communications_found("Найдено коммуникаций 1 из 1")

    with allure.step("Go to settings"):
        communications.click_settings()

    with allure.step("Change user to ecotelecom"):
        communications.go_to_user("Экотелеком")

    with allure.step("Go to communications"):
        communications.click_communications()

    with allure.step("Check that all communications found"):
        communications.assert_communications_found("Найдено коммуникаций 0 из 0")

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.e2e
@pytest.mark.calls
@allure.title("test_access_right_restt_for_user")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_access_right_restt_for_user")
def test_access_right_restt_for_user(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to page"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN_USER, PASSWORD)

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Click to actions with calls button (...)"):
        communications.press_calls_action_button_in_list(0)

    with allure.step("Check that no any stt buttons in menu"):
        expect(communications.menu).not_to_contain_text("Перераспознать")

    with allure.step("Close menu"):
        page.locator(BUTTON_CALLS_ACTION).nth(0).click()

    with allure.step("Click to action with one call button (...)"):
        communications.press_calls_action_button_in_list(1)

    with allure.step("Check that no any stt buttons in menu"):
        expect(communications.menu).not_to_contain_text("Перераспознать")

    with allure.step("Close menu"):
        page.locator(BUTTON_CALLS_ACTION).nth(1).click()

    with allure.step("Change access right restt: True"):
        rights = {"restt": True, "delete_call": False, "call_info_processing": False}
        give_access_right(API_URL, TOKEN_ADMIN, USER_ID_USER, rights)

    with allure.step("Reload page"):
        communications.reload_page()

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Click to actions with calls button (...)"):
        communications.press_calls_action_button_in_list(0)

    with allure.step("Check that stt buttons in menu"):
        expect(communications.menu).to_contain_text("Перераспознать")

    with allure.step("Close menu"):
        page.locator(BUTTON_CALLS_ACTION).nth(0).click()

    with allure.step("Click to action with one call"):
        communications.press_calls_action_button_in_list(1)

    with allure.step("Check that stt buttons in menu"):
        expect(communications.menu).to_contain_text("Перераспознать")

    with allure.step("Close menu"):
        page.locator(BUTTON_CALLS_ACTION).nth(1).click()

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.e2e
@pytest.mark.calls
@allure.title("test_access_right_delete_calls_for_user")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_access_right_delete_calls_for_user")
def test_access_right_delete_calls_for_user(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to page"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN_USER, PASSWORD)

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Click to actions with calls button (...)"):
        communications.press_calls_action_button_in_list(0)

    with allure.step("Check that no any delete buttons in menu"):
        expect(communications.menu).not_to_contain_text("Удалить коммуникации")

    with allure.step("Close menu"):
        page.locator(BUTTON_CALLS_ACTION).nth(0).click()

    with allure.step("Click to action with one call button (...)"):
        communications.press_calls_action_button_in_list(1)

    with allure.step("Check that no any delete buttons in menu"):
        expect(communications.menu).not_to_contain_text("Удалить коммуникацию")

    with allure.step("Close menu"):
        page.locator(BUTTON_CALLS_ACTION).nth(1).click()

    with allure.step("Change access_right delete_call: True"):
        rights = {"restt": False, "delete_call": True, "call_info_processing": False}
        give_access_right(API_URL, TOKEN_ADMIN, USER_ID_USER, rights)

    with allure.step("Reload page"):
        communications.reload_page()

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Click to actions with calls button (...)"):
        communications.press_calls_action_button_in_list(0)

    with allure.step("Check that delete button in menu"):
        expect(communications.menu).to_contain_text("Удалить коммуникации")

    with allure.step("Close menu"):
        page.locator(BUTTON_CALLS_ACTION).nth(0).click()

    with allure.step("Click to action with one call button (...)"):
        communications.press_calls_action_button_in_list(1)

    with allure.step("Check that delete button in menu"):
        expect(communications.menu).to_contain_text("Удалить коммуникацию")

    with allure.step("Close menu"):
        page.locator(BUTTON_CALLS_ACTION).nth(1).click()

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)

@pytest.mark.e2e
@pytest.mark.calls
@allure.title("test_access_right_processing_info_for_user")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_access_right_processing_info_for_user")
def test_access_right_processing_info_for_user(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to page"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN_USER, PASSWORD)

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Click to action with one call button (...)"):
        communications.press_calls_action_button_in_list(1)

    with allure.step("Open meta info"):
        communications.choose_from_menu_by_text_and_wait_for_modal("Мета инфо")

    with allure.step("Chek that no any processing info in meta info"):
        expect(communications.modal_window).not_to_contain_text("Время добавления звонка")
        expect(communications.modal_window).not_to_contain_text("Время завершения транскрибации")

    with allure.step("Close menu"):
        communications.button_cross.click()
        page.wait_for_selector(MODAL_WINDOW, state="hidden")

    with allure.step("Change access_right call_info_processing: True"):
        rights = {"restt": False, "delete_call": False, "call_info_processing": True}
        give_access_right(API_URL, TOKEN_ADMIN, USER_ID_USER, rights)

    with allure.step("Reload page"):
        communications.reload_page()

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Click to action with one call button (...)"):
        communications.press_calls_action_button_in_list(1)

    with allure.step("Open meta info"):
        communications.choose_from_menu_by_text_and_wait_for_modal("Мета инфо")

    with allure.step("Chek that processing info in meta info"):
        expect(communications.modal_window).to_contain_text("Время добавления звонка")
        #expect(page.locator(MODAL_WINDOW)).to_contain_text("Время завершения транскрибации")

    with allure.step("Close menu"):
        communications.close_modal_window()

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.e2e
@pytest.mark.calls
@allure.title("test_access_right_restt_for_manager")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_access_right_restt_for_manager")
def test_access_right_restt_for_manager(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create manager"):
        USER_ID_MANAGER, TOKEN_MANAGER, LOGIN_MANAGER = create_user(API_URL, ROLE_MANAGER, PASSWORD)

    with allure.step("Create user for manager"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)
        give_users_to_manager(API_URL, USER_ID_MANAGER, [USER_ID_USER, importFrom_user_id], TOKEN_MANAGER)

    with allure.step("Go to page"):
        communications.navigate(base_url)

    with allure.step("Auth with manager"):
        communications.auth(LOGIN_MANAGER, PASSWORD)

    with allure.step("Go to user"):
        communications.go_to_user(LOGIN_USER)

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Click to actions with calls button (...)"):
        communications.press_calls_action_button_in_list(0)

    with allure.step("Check that no any stt buttons in menu"):
        expect(communications.menu).not_to_contain_text("Перераспознать")

    with allure.step("Close menu"):
        page.locator(BUTTON_CALLS_ACTION).nth(0).click()

    with allure.step("Click to action with one call button (...)"):
        communications.press_calls_action_button_in_list(1)

    with allure.step("Check that no any stt buttons in menu"):
        expect(communications.menu).not_to_contain_text("Перераспознать")

    with allure.step("Close menu"):
        page.locator(BUTTON_CALLS_ACTION).nth(1).click()

    with allure.step("Change access right restt: True"):
        rights = {"restt":True,"delete_user":False,"add_user":False,"set_default_engine":False,"quota_edit":False,"gpt_quota":False,"user_modules_setup":False}
        give_access_right(API_URL, TOKEN_ADMIN, USER_ID_MANAGER, rights)

    with allure.step("Reload page"):
        communications.reload_page()

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Click to actions with calls button (...)"):
        communications.press_calls_action_button_in_list(0)

    with allure.step("Check that stt buttons in menu"):
        expect(communications.menu).to_contain_text("Перераспознать")

    with allure.step("Close menu"):
        page.locator(BUTTON_CALLS_ACTION).nth(0).click()

    with allure.step("Click to action with one call"):
        communications.press_calls_action_button_in_list(1)

    with allure.step("Check that stt buttons in menu"):
        expect(communications.menu).to_contain_text("Перераспознать")

    with allure.step("Close menu"):
        page.locator(BUTTON_CALLS_ACTION).nth(1).click()

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete manager"):
        delete_user(API_URL, TOKEN_MANAGER, USER_ID_MANAGER)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.e2e
@pytest.mark.calls
@allure.title("test_open_other_pages_from_communications")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_open_other_pages_from_communications. after https://task.imot.io/browse/DEV-3452")
def test_open_other_pages_from_communications(base_url, page: Page, context: BrowserContext) -> None:
    communications = Communications(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD, create_report=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        communications.go_to_user(LOGIN_USER)

    with allure.step("Open markup in new tab"):
        with context.expect_page() as new_tab_event:
            page.locator(BUTTON_MARKUP).click(modifiers=["Control"])
            new_tab=new_tab_event.value
            new_tab.wait_for_timeout(4000)

    with allure.step("Check"):
        expect(new_tab.get_by_text(LOGIN_USER, exact=True)).to_have_count(1, timeout=wait_until_visible)
        expect(new_tab.locator('[data-testid="markup_addGroup"]')).to_have_count(1, timeout=wait_until_visible)

    with allure.step("Close tab"):
        new_tab.close()

    with allure.step("Open notifications in new tab"):
        with context.expect_page() as new_tab_event:
            page.locator(BUTTON_NOTIFICATIONS).click(modifiers=["Control"])
            new_tab=new_tab_event.value
            new_tab.wait_for_timeout(4000)

    with allure.step("Check"):
        expect(new_tab.get_by_text(LOGIN_USER, exact=True)).to_have_count(1, timeout=wait_until_visible)
        expect(new_tab.locator('[class*="styles_addNewRule_"]')).to_have_count(1, timeout=wait_until_visible)

    with allure.step("Close tab"):
        new_tab.close()

    with allure.step("Open deals in new tab"):
        with context.expect_page() as new_tab_event:
            page.locator(BUTTON_DEALS).click(modifiers=["Control"])
            new_tab=new_tab_event.value
            new_tab.wait_for_timeout(4000)

    with allure.step("Check"):
        expect(new_tab.get_by_text(LOGIN_USER, exact=True)).to_have_count(1, timeout=wait_until_visible)
        expect(new_tab.locator('[data-testid="deals_btns_find"]')).to_have_count(1, timeout=wait_until_visible)

    with allure.step("Close tab"):
        new_tab.close()

    with allure.step("Open settings in new tab"):
        with context.expect_page() as new_tab_event:
            page.locator(BUTTON_SETTINGS).click(modifiers=["Control"])
            new_tab=new_tab_event.value
            new_tab.wait_for_timeout(4000)

    with allure.step("Check"):
        expect(new_tab.get_by_text(LOGIN_USER, exact=True)).to_have_count(3, timeout=wait_until_visible)
        expect(new_tab.locator(INPUT_NAME)).to_have_count(1, timeout=wait_until_visible)

    with allure.step("Close tab"):
        new_tab.close()

# reports
    with allure.step("Open create report in new tab"):
        communications.click_reports()
        with context.expect_page() as new_tab_event:
            page.locator('[href*="/report/create"]').click(modifiers=["Control"])
            new_tab=new_tab_event.value
            new_tab.wait_for_timeout(4000)

    with allure.step("Check"):
        expect(new_tab.get_by_text(LOGIN_USER, exact=True)).to_have_count(1, timeout=wait_until_visible)
        expect(new_tab.locator('[data-testid="reportMake"]')).to_have_count(1, timeout=wait_until_visible)

    with allure.step("Close tab"):
        new_tab.close()

    with allure.step("Open report management in new tab"):
        communications.click_reports()
        with context.expect_page() as new_tab_event:
            page.locator('[href*="/reports"]').click(modifiers=["Control"])
            new_tab=new_tab_event.value
            new_tab.wait_for_timeout(4000)

    with allure.step("Check"):
        expect(new_tab.get_by_text(LOGIN_USER, exact=True)).to_have_count(1, timeout=wait_until_visible)
        expect(new_tab.locator('[data-testid="addUserButton"]')).to_have_count(1, timeout=wait_until_visible)

    with allure.step("Close tab"):
        new_tab.close()

    with allure.step("Open report in new tab"):
        communications.click_reports()
        with context.expect_page() as new_tab_event:
            page.locator('[href*="/report/"]').nth(1).click(modifiers=["Control"])
            new_tab=new_tab_event.value
            new_tab.wait_for_timeout(4000)

    with allure.step("Check"):
        expect(new_tab.get_by_text(LOGIN_USER, exact=True)).to_have_count(1, timeout=wait_until_visible)
        expect(new_tab.locator('[data-testid="reportMake"]')).to_have_count(1, timeout=wait_until_visible)

    with allure.step("Close tab"):
        new_tab.close()
# reports

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_go_to_gpt_from_call")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_go_to_gpt_from_call")
def test_go_to_gpt_from_call(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN, PASSWORD)

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Go to Gpt from call"):
        page.locator(BUTTON_GPT).click()

    with allure.step("Check that opened"):
        expect(page.locator('[class*="styles_back"]')).to_have_count(1)
        expect(page.get_by_text("Применение GPT правила")).to_have_count(1)
        expect(page.get_by_text("Произвольный запрос")).to_have_count(1)
        expect(page.get_by_text("Дополнительные настройки")).to_have_count(1)
        expect(page.get_by_text("Системный текст")).to_have_count(1)
        expect(page.get_by_text("Движок")).to_have_count(1)
        expect(page.get_by_text("Модель")).to_have_count(1)
        expect(page.get_by_text("Температура")).to_have_count(1)
        expect(communications.button_accept).to_be_disabled()
        expect(page.locator(BUTTON_SAVE)).to_be_disabled()

    with allure.step("Fill question and assistant text"):
        page.locator(INPUT_GPT_QUESTION).type("12345", delay=10)

    with allure.step("Check that buttons enabled"):
        expect(communications.button_accept).to_be_enabled()
        expect(page.locator(BUTTON_SAVE)).to_be_enabled()

    with allure.step("Press (save to rules)"):
        page.wait_for_timeout(2000)
        page.locator(BUTTON_SAVE).click()
        page.wait_for_timeout(2000)

    with allure.step("Check that question have"):
        expect(page.locator(INPUT_GPT_QUESTION)).to_have_text("12345")

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_public_link_from_call_by_user")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_public_link_from_call_by_user")
def test_public_link_from_call_by_user(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN, PASSWORD)

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Click to (public link) button"):
        page.locator('[aria-label="Скопировать публичную ссылку"]').locator('[type="button"]').click()
        page.wait_for_timeout(1000)
        page.evaluate("""
                () => {
                    const input = document.createElement('input');
                    input.id = 'temp-clipboard-input';
                    document.body.appendChild(input);
                }
            """)
        page.focus("#temp-clipboard-input")
        page.keyboard.press("Control+V")
        copied_link = page.eval_on_selector("#temp-clipboard-input", "el => el.value")
        page.evaluate("document.getElementById('temp-clipboard-input').remove()")

        assert "&public=true" in copied_link

    with allure.step("go to private link"):
        communications.navigate(copied_link)

    with allure.step("check"):
        expect(page.locator('[aria-label="Скопировать публичную ссылку"]').locator('[type="button"]')).to_be_disabled()

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_public_link_from_call_by_admin")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_public_link_from_call_by_admin")
def test_public_link_from_call_by_admin(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with admin"):
        communications.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        communications.go_to_user(LOGIN_USER)

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Click to (public link) button"):
        page.locator('[aria-label="Скопировать публичную ссылку"]').locator('[type="button"]').click()
        page.wait_for_timeout(1000)
        page.evaluate("""
                () => {
                    const input = document.createElement('input');
                    input.id = 'temp-clipboard-input';
                    document.body.appendChild(input);
                }
            """)
        page.focus("#temp-clipboard-input")
        page.keyboard.press("Control+V")
        copied_link = page.eval_on_selector("#temp-clipboard-input", "el => el.value")
        page.evaluate("document.getElementById('temp-clipboard-input').remove()")

        assert "&public=true" in copied_link

    with allure.step("go to private link"):
        communications.navigate(copied_link)

    with allure.step("check"):
        expect(page.locator('[aria-label="Скопировать публичную ссылку"]').locator('[type="button"]')).to_be_disabled()

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_public_link_from_call_by_admin_to_not_logged_user")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_public_link_from_call_by_admin_to_not_logged_user")
def test_public_link_from_call_by_admin_to_not_logged_user(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with admin"):
        communications.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        communications.go_to_user(LOGIN_USER)

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Click to (public link) button"):
        page.locator('[aria-label="Скопировать публичную ссылку"]').locator('[type="button"]').click()
        page.wait_for_timeout(1000)
        page.evaluate("""
                () => {
                    const input = document.createElement('input');
                    input.id = 'temp-clipboard-input';
                    document.body.appendChild(input);
                }
            """)
        page.focus("#temp-clipboard-input")
        page.keyboard.press("Control+V")
        copied_link = page.eval_on_selector("#temp-clipboard-input", "el => el.value")
        page.evaluate("document.getElementById('temp-clipboard-input').remove()")

        assert "&public=true" in copied_link

    with allure.step("Quit from profile"):
        communications.quit_from_profile()

    with allure.step("go to private link"):
        communications.navigate(copied_link)

    with allure.step("check"):
        expect(page.locator(BUTTON_COMMUNICATIONS)).not_to_be_visible()
        expect(page.locator(BUTTON_REPORTS)).not_to_be_visible()
        expect(page.locator(BUTTON_MARKUP)).not_to_be_visible()
        expect(page.locator(BUTTON_NOTIFICATIONS)).not_to_be_visible()
        expect(page.locator(BUTTON_DEALS)).not_to_be_visible()
        expect(page.locator(BUTTON_SETTINGS)).not_to_be_visible()
        expect(page.locator(AUDIO_PLAYER)).to_be_visible()
        #expect(page.locator(BUTTON_CALLS_ACTION)).to_be_disabled()
        expect(page.locator(BUTTON_EXPAND_CALL)).not_to_be_visible()
        expect(page.locator(BUTTON_GPT)).not_to_be_visible()
        expect(page.locator(BUTTON_SHARE_CALL)).not_to_be_visible()
        expect(page.locator(BUTTON_ADD_COMMENT)).not_to_be_visible()
        expect(page.get_by_text("0987654321")).not_to_be_visible()
        expect(page.get_by_text("Теги коммуникации")).to_be_visible()
        expect(page.get_by_text("00:00:38")).to_be_visible()
        expect(page.get_by_text("1234567890")).to_be_visible()
        #expect(page.get_by_text("23.04.25 10:16")).to_be_visible()

        # expect(page.locator(BUTTON_SHARE_CALL)).not_to_be_visible()
        # expect(page.locator(BUTTON_SHARE_CALL)).not_to_be_visible()

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_communication_check_list_in_open_call")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_communication_check_list_in_open_call")
def test_communication_check_list_in_open_call(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN, PASSWORD)

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Check"):
        expect(page.locator(CHECK_LIST_VALUES_IN_CALL_HEADER)).to_have_text("50%0 баллов")
        expect(page.locator(BLOCK_CHECK_LISTS).locator(CHECK_LIST_PERCENT)).to_have_text("50%")
        expect(page.locator(BLOCK_CHECK_LISTS).locator(POINTS_IN_CHECK_LISTS_LIST)).to_have_text("0 баллов")
        expect(page.locator(BLOCK_CHECK_LISTS).locator(CHECK_LIST_NAME)).to_have_text("auto_call_ch_list")
        expect(page.locator(TOOLTIP_IN_CHECK_LIST)).to_have_count(1)
        expect(page.locator(BLOCK_CHECK_LISTS).locator(CHECK_LIST_MIN_MAX)).to_have_text("мин: -10макс: 10")

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_deal_check_list_in_open_call")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_deal_check_list_in_open_call")
def test_deal_check_list_in_open_call(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with Ecotelecom"):
        communications.auth(ECOTELECOM, ECOPASS)

    with allure.step("Choose period from 01/01/2022 to 31/12/2022"):
        communications.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Fill ID to find call"):
        communications.fill_id("1644268692.90190")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_less_than_50()

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Check"):
        expect(page.locator(CHECK_LIST_VALUES_IN_CALL_HEADER)).to_have_text("59%27 баллов")
        expect(page.locator(BLOCK_CHECK_LISTS).locator(CHECK_LIST_PERCENT)).to_have_text("59%")
        expect(page.locator(BLOCK_CHECK_LISTS).locator(POINTS_IN_CHECK_LISTS_LIST)).to_have_text("27 баллов")
        expect(page.locator(BLOCK_CHECK_LISTS).locator(CHECK_LIST_NAME)).to_have_text("Тестовый чеклист сделки")
        expect(page.locator(TOOLTIP_IN_CHECK_LIST)).to_have_count(1)
        expect(page.locator(BLOCK_CHECK_LISTS).locator(CHECK_LIST_MIN_MAX)).to_have_text("мин: 0макс: 46")

@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_communication_and_deal_check_list_in_open_call")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_communication_and_deal_check_list_in_open_call")
def test_communication_and_deal_check_list_in_open_call(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with Ecotelecom"):
        communications.auth(ECOTELECOM, ECOPASS)

    with allure.step("Choose period from 01/01/2022 to 31/12/2022"):
        communications.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Fill ID to find call"):
        communications.fill_id("1644298753.90325")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_less_than_50()

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Check"):
        # check percent and points for two check-lists
        expect(page.locator(CHECK_LIST_VALUES_IN_CALL_HEADER)).to_have_text("100%50 баллов59%27 баллов")
        # check first check-list
        expect(page.locator(BLOCK_CHECK_LISTS).locator(CHECK_LIST_NAME).nth(0)).to_have_text("Чеклист только с при приоритетной оценкой")
        expect(page.locator(BLOCK_CHECK_LISTS).locator(CHECK_LIST_PERCENT).nth(0)).to_have_text("100%")
        expect(page.locator(BLOCK_CHECK_LISTS).locator(POINTS_IN_CHECK_LISTS_LIST).nth(0)).to_have_text("50 баллов")
        expect(page.locator(BLOCK_CHECK_LISTS).locator(CHECK_LIST_MIN_MAX).nth(0)).to_have_text("мин: 0макс: 50")
        expect(page.locator(BLOCK_CHECK_LISTS).locator('[class="ChekListAppraised"]')).to_have_text("Сработала приоритетная оценка")
        # check second check-list
        expect(page.locator(BLOCK_CHECK_LISTS).locator(CHECK_LIST_NAME).nth(1)).to_have_text("Тестовый чеклист сделки")
        expect(page.locator(BLOCK_CHECK_LISTS).locator(CHECK_LIST_PERCENT).nth(1)).to_have_text("59%")
        expect(page.locator(BLOCK_CHECK_LISTS).locator(POINTS_IN_CHECK_LISTS_LIST).nth(1)).to_have_text("27 баллов")
        expect(page.locator(BLOCK_CHECK_LISTS).locator(CHECK_LIST_MIN_MAX).nth(1)).to_have_text("мин: 0макс: 46")
        # check (?) for both check-list
        expect(page.locator(TOOLTIP_IN_CHECK_LIST)).to_have_count(2)

@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_calls_actions_apply_gpt_if_500")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_calls_actions_apply_gpt_if_500")
def test_calls_actions_apply_gpt_if_500(base_url, page: Page) -> None:
    communications = Communications(page)

    def handle_gpt(route: Route):
        route.fulfill(status=500, body="")
    # Intercept the route
    page.route("**/gpt/", handle_gpt)

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD, gpt_rule=True, upload_call=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN, PASSWORD)

    with allure.step("Press button (Calls action)"):
        communications.press_calls_action_button_in_list(0)

    with allure.step("Choose Apply gpt"):
        communications.choose_from_menu_by_text_and_wait_for_modal("Применить GPT")

    with allure.step("Check alert"):
        communications.check_alert("Ошибка, попробуйте позднее")

    with allure.step("Check modal"):
        expect(page.locator(MODAL_WINDOW).locator(CHECKBOX)).to_have_count(0)
        expect(page.locator(MODAL_WINDOW).locator(BUTTON_ADD_GPT_RULE)).to_have_count(0)
        expect(page.locator(MODAL_WINDOW).locator(BUTTON_ACCEPT)).to_be_disabled()
        expect(page.locator(MODAL_WINDOW).locator(BUTTON_OTMENA)).to_be_enabled()
        expect(page.locator(MODAL_WINDOW).locator('[class*="_empty__title_"]')).to_have_text("У вас пока нет правил GPT")

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)

@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_calls_actions_apply_gpt_without_gpt_rule")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_calls_actions_apply_gpt_without_gpt_rule")
def test_calls_actions_apply_gpt_without_gpt_rule(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN, PASSWORD)

    with allure.step("Press button (Calls action)"):
        communications.press_calls_action_button_in_list(0)

    with allure.step("Choose Apply gpt"):
        communications.choose_from_menu_by_text_and_wait_for_modal("Применить GPT")

    with allure.step("Check modal"):
        expect(page.locator(MODAL_WINDOW).locator(CHECKBOX)).to_have_count(0)
        expect(page.locator(MODAL_WINDOW).locator(BUTTON_ADD_GPT_RULE)).to_have_count(0)
        expect(page.locator(MODAL_WINDOW).locator(BUTTON_ACCEPT)).to_be_disabled()
        expect(page.locator(MODAL_WINDOW).locator(BUTTON_OTMENA)).to_be_enabled()
        expect(page.locator(MODAL_WINDOW).locator('[class*="_empty__title_"]')).to_have_text("У вас пока нет правил GPT")

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_calls_actions_apply_gpt")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_calls_actions_apply_gpt. all rules and one rule")
def test_calls_actions_apply_gpt(base_url, page: Page) -> None:
    communications = Communications(page)

    warn = "Количество коммуникаций, к которым применятся все активные правила GPT"
    count = "1 шт"

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD, gpt_rule=True, upload_call=True)

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN, PASSWORD)

    with allure.step("Press button (Calls action)"):
        communications.press_calls_action_button_in_list(0)

    with allure.step("Choose Apply gpt"):
        communications.choose_from_menu_by_text_and_wait_for_modal("Применить GPT")

    with allure.step("Check modal"):
        expect(communications.modal_window.locator(CHECKBOX)).not_to_be_checked()
        expect(communications.modal_window.locator(BUTTON_ADD_GPT_RULE)).to_be_disabled()
        expect(communications.modal_window.locator(BUTTON_ACCEPT)).to_be_enabled()
        expect(communications.modal_window.locator(BUTTON_OTMENA)).to_be_enabled()
        expect(communications.modal_window.locator('[class=" css-hlgwow"]')).to_have_text("Все правила")

    with allure.step("Press (accept) button"):
        communications.button_accept.click()

    with allure.step("Check warning in modal"):
        expect(communications.modal_window.locator('[class*="styles_contentSubmit__title_"]')).to_have_text(warn)
        expect(communications.modal_window.locator('[class*="styles_contentSubmit__count_"]')).to_have_text(count)

    with allure.step("press (cancel) and go to initial screen"):
        page.locator(BUTTON_OTMENA).click()

    with allure.step("Check modal"):
        expect(communications.modal_window.locator(CHECKBOX)).not_to_be_checked()
        expect(communications.modal_window.locator(BUTTON_ADD_GPT_RULE)).to_be_disabled()
        expect(communications.modal_window.locator(BUTTON_ACCEPT)).to_be_enabled()
        expect(communications.modal_window.locator(BUTTON_OTMENA)).to_be_enabled()
        expect(communications.modal_window.locator('[class=" css-hlgwow"]')).to_have_text("Все правила")

    with allure.step("Press (accept) button"):
        communications.button_accept.click()

    with allure.step("Check warning in modal"):
        expect(communications.modal_window.locator('[class*="styles_contentSubmit__title_"]')).to_have_text(warn)
        expect(communications.modal_window.locator('[class*="styles_contentSubmit__count_"]')).to_have_text(count)

    with allure.step("AGAIN press (accept) button"):
        communications.button_accept.click()

    with allure.step("Check alert"):
        communications.check_alert("Действие начато")

    with allure.step("Press button (Calls action)"):
        communications.press_calls_action_button_in_list(0)

    with allure.step("Choose Apply gpt"):
        communications.choose_from_menu_by_text_and_wait_for_modal("Применить GPT")

    with allure.step("Choose rule"):
        communications.modal_window.locator('[viewBox="0 0 20 20"]').click()
        page.wait_for_selector(MENU)
        communications.menu.get_by_text("auto_gpt_rule", exact=True).click()

    with allure.step("Check modal"):
        expect(communications.modal_window.locator('[class=" css-hlgwow"]')).to_have_text("auto_gpt_rule")
        expect(communications.modal_window.locator(BUTTON_ADD_GPT_RULE)).to_be_enabled()
        expect(communications.modal_window.locator(BUTTON_ACCEPT)).to_be_enabled()
        expect(communications.modal_window.locator(BUTTON_OTMENA)).to_be_enabled()

    with allure.step("Add additional select"):
        page.locator(BUTTON_ADD_GPT_RULE).click()
        page.wait_for_selector('[class*="_additionalRuleSelect_"]')

    with allure.step("Delete additional select"):
        page.locator('[data-testid="deleteRuleBtn"]').click()
        page.wait_for_selector('[class*="_additionalRuleSelect_"]', state="hidden")

    with allure.step("AGAIN press (accept) button"):
        communications.button_accept.click()

    with allure.step("Check alert"):
        communications.check_alert("Действие начато")

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_translation_in_communication")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_check_translation_in_communication. Also have api tests")
def test_check_translation_in_communication(base_url, page: Page) -> None:
    communications = Communications(page)

    requests = []

    with allure.step("Go to url"):
        communications.navigate("http://192.168.10.101/feature-dev-3474/")

    with allure.step("Auth with Ecotelecom"):
        communications.auth(ECOTELECOM, ECOPASS)

    with allure.step("Choose period from 01/01/2022 to 31/12/2022"):
        communications.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Fill ID to find call"):
        communications.fill_id("1644268426.90181")

    with allure.step("Press button (Find communications)"):
        communications.press_find_communications_less_than_50()

    with allure.step("Expand call"):
        communications.expand_call()

    with allure.step("Press tranlsation button"):
        page.get_by_text("Перевод").click()
        page.wait_for_selector(MENU)

    with allure.step("Choose first language"):
        communications.choose_option(1)

    with allure.step("Check"):
        expect(page.locator('[class*="styles_topTitleRight"]')).to_have_text("Переведенный текст")
        expect(page.locator('[class*="styles_original_"]')).to_have_count(7)
        expect(page.locator('[class*="styles_translated_"]')).to_have_count(7)

    with allure.step("Press tranlsation button"):
        page.get_by_text("Bulgarian").click()
        page.wait_for_selector(MENU)

    with allure.step("Request capture start"):
        page.on("request", lambda request: (
            requests.append({
                'url': request.url,
                'method': request.method
            }) if request.resource_type in ['xhr', 'fetch'] else None
        ))

    with allure.step("Choose original language"):
        communications.choose_option(0)

    with allure.step("Check requests list. When original - no any requests"):
        page.wait_for_timeout(5000)
        assert len(requests) == 0

    with allure.step("Check"):
        expect(page.locator('[class*="styles_topTitleRight"]')).to_have_count(0)
        expect(page.locator('[class*="styles_original_"]')).to_have_count(0)
        expect(page.locator('[class*="styles_translated_"]')).to_have_count(0)


@pytest.mark.calls
@pytest.mark.e2e
@allure.title("test_check_tags_more")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_check_tags_more")
def test_check_tags_more(base_url, page: Page) -> None:
    communications = Communications(page)

    with allure.step("Create user"):
        USER_ID_USER1, TOKEN_USER1, LOGIN_USER1 = create_user(API_URL, ROLE_USER, PASSWORD,
                                                              upload_call=True,
                                                              create_many_rules=True,
                                                              rules_entity="DEAL",
                                                              rules_amount=11,
                                                              rule_value="deal"
                                                              )
        USER_ID_USER2, TOKEN_USER2, LOGIN_USER2 = create_user(API_URL, ROLE_USER, PASSWORD,
                                                              upload_call=True,
                                                              create_many_rules=True,
                                                              phrases_and_dicts=["один"],
                                                              rules_amount=11,
                                                              rule_value="fragment"
                                                              )

    with allure.step("Go to url"):
        communications.navigate(base_url)

    with allure.step("Auth with user"):
        communications.auth(LOGIN_USER1, PASSWORD)

    with allure.step("Check alll tags"):
        expect(communications.deal_tag).to_have_count(10)
        expect(communications.block_one_communication.get_by_text("Показать еще (1)")).to_have_count(1)
        expect(communications.communication_tag).to_have_count(10)
        expect(communications.block_one_communication.get_by_text("Показать еще (2)")).to_have_count(1)
        expect(communications.block_one_communication.get_by_text("Скрыть")).to_have_count(0)
        expect(communications.block_one_communication.get_by_text("Теги сделки")).to_have_count(1)
        expect(communications.block_one_communication.get_by_text("Теги коммуникации")).to_have_count(1)

    with allure.step("Open deal tags"):
        page.get_by_text("Показать еще (1)").click()

    with allure.step("Check that deal tags opened"):
        expect(communications.deal_tag).to_have_count(11)
        expect(communications.block_one_communication.get_by_text("Показать еще (1)")).to_have_count(0)
        expect(communications.block_one_communication.get_by_text("Скрыть")).to_have_count(1)

    with allure.step("Close deal tags"):
        page.get_by_text("Скрыть").click()

    with allure.step("Check that deal tags closed"):
        expect(communications.deal_tag).to_have_count(10)
        expect(communications.block_one_communication.get_by_text("Показать еще (1)")).to_have_count(1)
        expect(communications.block_one_communication.get_by_text("Скрыть")).to_have_count(0)

    with allure.step("Open communication tags"):
        page.get_by_text("Показать еще (2)").click()

    with allure.step("Check that communication tags opened"):
        expect(communications.communication_tag).to_have_count(12)
        expect(communications.block_one_communication.get_by_text("Показать еще (2)")).to_have_count(0)
        expect(communications.block_one_communication.get_by_text("Скрыть")).to_have_count(1)

    with allure.step("Close deal tags"):
        page.get_by_text("Скрыть").click()

    with allure.step("Check that deal tags closed"):
        expect(communications.communication_tag).to_have_count(10)
        expect(communications.block_one_communication.get_by_text("Показать еще (2)")).to_have_count(1)
        expect(communications.block_one_communication.get_by_text("Скрыть")).to_have_count(0)

    with allure.step("Quit from user1"):
        communications.quit_from_profile()

    with allure.step("Auth with user"):
        communications.auth(LOGIN_USER2, PASSWORD)

    with allure.step("Check alll tags"):
        expect(communications.fragment_tag).to_have_count(10)
        expect(communications.block_one_communication.get_by_text("Показать еще 1")).to_have_count(1)
        expect(communications.communication_tag).to_have_count(10)
        expect(communications.block_one_communication.get_by_text("Показать еще (2)")).to_have_count(1)
        expect(communications.block_one_communication.get_by_text("Скрыть")).to_have_count(0)
        expect(communications.block_one_communication.get_by_text("Теги фрагментов")).to_have_count(1)
        expect(communications.block_one_communication.get_by_text("Теги коммуникации")).to_have_count(1)

    with allure.step("Open fragment tags"):
        page.get_by_text("Показать еще 1").click()

    with allure.step("Check that fragment tags opened"):
        expect(communications.fragment_tag).to_have_count(11)
        expect(communications.block_one_communication.get_by_text("Показать еще 1")).to_have_count(0)
        expect(communications.block_one_communication.get_by_text("Скрыть")).to_have_count(1)

    with allure.step("Close fragment tags"):
        page.get_by_text("Скрыть").click()

    with allure.step("Check that fragment tags closed"):
        expect(communications.fragment_tag).to_have_count(10)
        expect(communications.block_one_communication.get_by_text("Показать еще 1")).to_have_count(1)
        expect(communications.block_one_communication.get_by_text("Скрыть")).to_have_count(0)

    with allure.step("Open communication tags"):
        page.get_by_text("Показать еще (2)").click()

    with allure.step("Check that communication tags opened"):
        expect(communications.communication_tag).to_have_count(12)
        expect(communications.block_one_communication.get_by_text("Показать еще (2)")).to_have_count(0)
        expect(communications.block_one_communication.get_by_text("Скрыть")).to_have_count(1)

    with allure.step("Close deal tags"):
        page.get_by_text("Скрыть").click()

    with allure.step("Check that deal tags closed"):
        expect(communications.communication_tag).to_have_count(10)
        expect(communications.block_one_communication.get_by_text("Показать еще (2)")).to_have_count(1)
        expect(communications.block_one_communication.get_by_text("Скрыть")).to_have_count(0)


    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER1, USER_ID_USER1)
        delete_user(API_URL, TOKEN_USER2, USER_ID_USER2)